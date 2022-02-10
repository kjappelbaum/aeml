# -*- coding: utf-8 -*-
import datetime
import os
from typing import Tuple, Union

import matplotlib.pyplot as plt
import openpyxl
import pandas as pd
from dateutil.parser import parse


def get_stepchange_times(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["WW Data"]

    date = parse(str(ws.cell(1, 2).value))
    start = parse(str(ws.cell(4, 2).value))
    end = parse(str(ws.cell(5, 2).value))

    return {
        "start": datetime.datetime(date.year, date.month, date.day, start.hour, start.minute),
        "end": datetime.datetime(date.year, date.month, date.day, end.hour, end.minute),
    }


def load_process_data(
    filepath: Union[str, bytes, os.PathLike]
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    df_measurements = pd.read_excel(
        filepath,
        sheet_name="Process Data",
        skiprows=[0, 1, 2, 3, 5, 6],
        engine="openpyxl",
    )
    times = get_stepchange_times(filepath)
    if "valve-position-12" not in df_measurements.columns:
        print("valve NOT in columns")
        df_measurements["valve-position-12"] = [1] * len(df_measurements)
    else:
        print("valve in columns")
    ti1213 = []
    for _, row in df_measurements.iterrows():
        if row["valve-position-12"] == 1:
            ti1213.append(row["TI-12"])
        else:
            ti1213.append(row["TI-13"])

    df_measurements["TI-1213"] = ti1213
    # Let's do something dirty, but Pythonic
    try:
        # one of the first days
        df_gas = pd.read_excel(filepath, sheet_name="GasMET2", engine="openpyxl")
        df_gas["Time"] = [v.split()[-1] for v in df_gas["Time"].astype(str).values]

        df_gas["Datetime"] = pd.to_datetime(
            df_gas["Date"].astype(str) + " " + df_gas["Time"], dayfirst=True
        )

        df_gas["Datetime"].tz_localize = None

    except Exception as e:
        # One of the later days
        print(filepath, e)
        df_gas = pd.read_excel(filepath, sheet_name="GasMET2", skiprows=[1], engine="openpyxl")
        df_gas["Datetime"] = pd.to_datetime(df_gas["Time"].astype(str).values)

    df_gas = df_gas.set_index(df_gas["Datetime"])

    df_measurements = df_measurements.set_index(pd.to_datetime(df_measurements["Date"].values))

    try:
        df_measurements.index = df_measurements.index.tz_convert(None)
    except Exception:
        pass

    try:
        df_gas.index = df_gas.index.tz_convert(None)
    except Exception:
        pass

    df_measurements.index.name = None
    df_gas.index.name = None

    return df_measurements, df_gas, times


